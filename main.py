from selenium.webdriver.chrome.webdriver import WebDriver
from selenium.common.exceptions import NoSuchElementException
import time
import re
from openpyxl import load_workbook

fn = 'dane.xlsx'
wb = load_workbook(filename=fn)
ws = wb.active

d = WebDriver()

i = 31
all_links = []

while i < 164:
    d.get(f'https://rejestradwokatow.pl/adwokat/wyszukaj/strona/{i}/pra/1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,23,24,25,26,27/sta/2')

    links = d.find_elements_by_css_selector(".icon_link [href]")
    for link in links:
        href = link.get_attribute("href")
        all_links.append(href)
    i += 1
    time.sleep(0.2)

print(len(all_links))

lp = 3001
x = 3002
for link in all_links:
    d.get(link)
    section = d.find_element_by_tag_name("section")
    name = section.find_element_by_tag_name("h2").text
    info = section.find_element_by_xpath("//div[@class='line_list_K']")


    # szukaj emaila
    email = ''
    try:
        email = info.find_element_by_xpath("//div[@class='address_e']")
        email_1 = email.get_attribute("data-ea")
        email_2 = email.get_attribute("data-eb")
        if email_1:
            email = email_1 + '@' + email_2
    except NoSuchElementException:
        pass


    #   szukaj telefonu
    number = ''
    try:
        number = info.find_element_by_xpath("//div/span[text()='Komórkowy:']").find_element_by_xpath('..').find_element_by_tag_name("div").text
    except NoSuchElementException:
        try:
            number = info.find_element_by_xpath("//div/span[text()='Stacjonarny:']").find_element_by_xpath('..').find_element_by_tag_name("div").text
        except NoSuchElementException:
            print('szukam w drugiej czesci...')
            phone = d.find_element_by_xpath("//div[@class='mb_tab_content special_one']/div[@class='line_list_K']").text
            try:
                number = re.search(r'(?<!\w)(\(?(\+|00)?48\)?)?[ -]?\d{3}[ -]?\d{3}[ -]?\d{3}(?!\w)|(\d{10})|(\+58\d{9})', phone).group(0)
            except AttributeError:
                pass

    #   szukaj adresu
    address = ''
    try:
        address = info.find_element_by_xpath("//div/span[text()='Adres do korespondencji:']").find_element_by_xpath('..').find_element_by_tag_name("div").text
    except NoSuchElementException:
        pass
    else:
        adres = d.find_element_by_xpath("//div[@class='mb_tab_content special_one']/div[@class='line_list_K']").text
        try:
            if address == '':
                # address = re.search(r"^(\d{1}|\d{2}|[uU]l\.|[uU]lica|[pP]l\.|[Pp]lac|[Aa]l\.|[Aa]lej[ea]|[Rr]ondo|[Oo]siedle|[Oo]s)[^\n]*",
                #                     adres, re.MULTILINE).group(0) + ', '
                address = re.search(r"^[kK]ancelaria [aA][a-zA-Z](.*(\n|\r|\r\n)){3}", adres, re.MULTILINE).group(0)

                # address += re.search(r'[0-9]{2}-[0-9]{3} [a-zA-ZąćęłńóśźżĄĘŁŃÓŚŹŻ ]*', adres).group(0)
                # address = ulica + '\n'
                # print(adres)
                # post_code = re.search(r'([0-9]{2}-[0-9]{3})|([0-9]{5}) [a-zA-ZąćęłńóśźżĄĘŁŃÓŚŹŻ ]*', adres).group(0)
                # address = post_code + '\n'

        except Exception as e:
            print(e)

    try:
        name = name.strip()
        email = email.strip()
        number = number.strip()
        address = address.strip()
        print("Imię, nazwisko: %s" % name)
        print("Email: %s" % email)
        print("Telefon: %s" % number)
        print("Adres: %s" % address)

        print(lp)
    except AttributeError:
        pass

    try:
        ws[f'A{x}'] = lp
        ws[f'B{x}'] = name
        ws[f'C{x}'] = email
        ws[f'D{x}'] = number
        ws[f'E{x}'] = address
        wb.save(fn)
    except Exception as e:
        print(e)

    x += 1
    lp += 1


# print(address.text)
